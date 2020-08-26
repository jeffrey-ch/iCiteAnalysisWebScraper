#!/usr/bin/env python
# coding: utf-8

# In[1]:


# Import necessary libraries

from selenium import webdriver
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import StaleElementReferenceException


# In[3]:


# List of relevant names imported from an Excel file

# Insert the name of the XCEL list for the load_workbook
wb = load_workbook("Test Excel.xlsx")
# Insert sheet name for the wb[" "] 
ws = wb["Sheet1"]
# Inser the name column for the ws[" "] 
column = ws["A"]
names = [column[x].value for x in range(len(column))]


# In[4]:


# Initialize variables for relevant variables needed
rcr_values = []
pub_year = []
avg_hum = []

# Numerizes the maximum number of names, used later to break the program
length = len(names)-1

# Url of the icite website
url = "https://icite.od.nih.gov/analysis"

# Initializes the chromedriver based on designated url path in (r"chromedriver.exe location")
browser = webdriver.Chrome(r" ")

# Opens the icite website in the chrome driver
browser.get(url)

# Initializes the searching for all the names in the list
for x in names:
    name = x
    
    # Input the instance of the name in the name field and selects process
    name_input = "/html/body/div/div/form/div[2]/input"
    submit_input = "/html/body/div/div/form/div[6]/button/span"
    browser.find_element_by_xpath(name_input).send_keys(name)
    time.sleep(0.5)
    d_e = browser.find_element_by_xpath(submit_input)
    while True:
        if d_e.is_enabled() == False:
            continue
        else:
            browser.find_element_by_xpath(submit_input).click()
            break
    
    # Makes sure loop continues if network or other issues slows the program
    while True: 
        # Try except case accounts for whether the immediate results is successful into getting the analysis page
        try:
            
            # The UI in the analysis page is found
            found = browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/ul")

            # Creates empty list for the names to be broken up and cross validated
            i_names = []
            name_m = [name]

            # Determines the author names of the first relevant paper
            value = browser.find_element_by_xpath("/html/body/div[1]/div/div[7]/div/div/div[1]/div/div[3]/div[2]/div/div/div[1]/div[3]")
            i_names.append(value.text)

            # Strips authors name into first name and last names
            strip_i_names = [n for ns in i_names for n in ns.split(",")]
            strip_i2_names = [n1 for n1s in strip_i_names for n1 in n1s.split(" ")]
            strip_name = [n2 for n2s in name_m for n2 in n2s.split(" ")]

            # If both the first and last name of the relevant name is checked to determine if it is within the author names of the first paper
            if strip_name[0] and strip_name[1] in strip_i2_names:

                # Identifies the location of mean rcr, pubs/year, and average human score
                rcr_mean_loc = "/html/body/div[1]/div/div[3]/div[1]/div[1]/table/tbody/tr/td[9]"
                rcr_mean = browser.find_element_by_xpath(rcr_mean_loc)
                rcr_values.append(rcr_mean.text)

                browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/ul/li[2]/a").click()
                time.sleep(1)

                pub_year_loc = "/html/body/div[1]/div/div[3]/div[2]/div[1]/table/tbody/tr/td[2]"
                pub_year_mean = browser.find_element_by_xpath(pub_year_loc)
                pub_year.append(pub_year_mean.text)

                avg_hum_loc = "/html/body/div[1]/div/div[3]/div[2]/div[1]/table/tbody/tr/td[3]"
                avg_hum_mean = browser.find_element_by_xpath(avg_hum_loc)
                avg_hum.append(avg_hum_mean.text)

                # If the index of the name reaches the end of the list, the loop is broken, if not, process starts over with next name
                if names.index(x) == length:
                    break
                else:
                    browser.execute_script("window.history.go(-1)")
                    time.sleep(1)
                    browser.find_element_by_xpath(name_input).clear()
                    break

            # If cross check does not find the matching name, process starts over with next name, and gives 0 to the relevant variables 
            else:
                rcr_values.append("0")
                pub_year.append("0")
                avg_hum.append("0")
                browser.execute_script("window.history.go(-1)")
                time.sleep(1)
                browser.find_element_by_xpath(name_input).clear()
                break

        # If next page fails, meaning either no results or too many results
        except NoSuchElementException:
            
            try: 
                # If no results, breaks out of analysis-checking repeat loop
                if "No results found." == browser.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div/p").text:
                    break
                
                # If author exists but no pubmed, breaks out of analysis-checking repeat loop
                elif "Submission did not contain any PubMed IDs that we have data for." == browser.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div/p").text:
                    break
                
                # If search fails, researches until successful and continues on to analysis portion
                elif "Search failed." == browser.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div/p").text:
                    time.sleep(0.01)
                    browser.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div/div/button").click()
                    while True:
                        if d_e.is_enabled() == False:
                            continue
                        else:
                            browser.find_element_by_xpath(submit_input).click()
                            break
                    continue
                    
                # If too many results, clicks OK to get to analysis page
                else:
                    time.sleep(1)
                    browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/div/div/div/form/button[2]").click()

                    # Creates empty list for the names to be broken up and cross validated
                    i_names = []
                    name_m = [name]

                    # Determines the author names of the first relevant paper
                    value = browser.find_element_by_xpath("/html/body/div[1]/div/div[7]/div/div/div[1]/div/div[3]/div[2]/div/div/div[1]/div[3]")
                    i_names.append(value.text)

                    # Strips authors name into first name and last names
                    strip_i_names = [n for ns in i_names for n in ns.split(",")]
                    strip_i2_names = [n1 for n1s in strip_i_names for n1 in n1s.split(" ")]
                    strip_name = [n2 for n2s in name_m for n2 in n2s.split(" ")]

                    # If both the first and last name of the relevant name is checked to determine if it is within the author names of the first paper
                    if strip_name[0] and strip_name[1] in strip_i2_names:

                        # Identifies the location of mean rcr, pubs/year, and average human score
                        rcr_mean_loc = "/html/body/div[1]/div/div[3]/div[1]/div[1]/table/tbody/tr/td[9]"
                        rcr_mean = browser.find_element_by_xpath(rcr_mean_loc)
                        rcr_values.append(rcr_mean.text)

                        browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/ul/li[2]/a").click()
                        time.sleep(1)

                        pub_year_loc = "/html/body/div[1]/div/div[3]/div[2]/div[1]/table/tbody/tr/td[2]"
                        pub_year_mean = browser.find_element_by_xpath(pub_year_loc)
                        pub_year.append(pub_year_mean.text)

                        avg_hum_loc = "/html/body/div[1]/div/div[3]/div[2]/div[1]/table/tbody/tr/td[3]"
                        avg_hum_mean = browser.find_element_by_xpath(avg_hum_loc)
                        avg_hum.append(avg_hum_mean.text)

                        # If the index of the name reaches the end of the list, the loop is broken, if not, process starts over with next name
                        if names.index(x) == length:
                            break
                        else:
                            browser.execute_script("window.history.go(-1)")
                            time.sleep(1)
                            browser.find_element_by_xpath(name_input).clear()
                            break

                    # If cross check does not find the matching name, process starts over with next name, and gives 0 to the relevant variables
                    else:
                        rcr_values.append("0")
                        pub_year.append("0")
                        avg_hum.append("0")
                        browser.execute_script("window.history.go(-1)")
                        time.sleep(1)
                        browser.find_element_by_xpath(name_input).clear()
                        break
            
            # If initial analysis page failed to load, exceptions carry back to try
            except NoSuchElementException:
                continue
            
            # If an interactable element fails, exceptions carry back to try
            except ElementNotInteractableException:
                continue
            
            # If element is removed from document, exceptions carry back to try
            except StaleElementReferenceException:
                continue
            
    # If no results, click OK and assigns 0 values to relevant variables and continues to next name
    if "No results found." == browser.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div/p").text:
                time.sleep(0.01)
                browser.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div/div/button").click()
                browser.find_element_by_xpath(name_input).clear()
                rcr_values.append("0")
                pub_year.append("0")
                avg_hum.append("0")
                time.sleep(1)
                continue
                
    # If no submissions, click OK and assigns 0 values to relevant variables and continues to next name
    elif "Submission did not contain any PubMed IDs that we have data for." == browser.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div/p").text:
                time.sleep(0.01)
                browser.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/div/div/div/button").click()
                browser.find_element_by_xpath(name_input).clear()
                rcr_values.append("0")
                pub_year.append("0")
                avg_hum.append("0")
                browser.execute_script("window.history.go(-1)")
                browser.find_element_by_xpath(name_input).clear()
                time.sleep(1)
                continue
            


# In[6]:


# Creates a data frame with names, rcr_values, pubs/year, and average human score and exports it to an exile file 

df = pd.DataFrame()
df["Names"] = names
df["Mean RCRs"] = rcr_values
df["Pubs Per Year"] = pub_year
df["Average Human Score"] = avg_hum

# Insert name of the excel file you want to export to 
df.to_excel("Excel Results.xlsx", index=False)

